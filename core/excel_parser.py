"""
Excel 檔案解析功能
"""
import os
import time
import zipfile
import xml.etree.ElementTree as ET
import re
import json
import hashlib
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.worksheet.formula import ArrayFormula
import config.settings as settings
from utils.cache import copy_to_cache
import logging
import urllib.parse

def extract_external_refs(xlsx_path):
    """
    解析 Excel xlsx 中 external reference mapping: [n] -> 路徑
    支援兩種來源：
    - xl/externalLinks/externalLinkN.xml 的 externalBookPr@href
    - xl/externalLinks/_rels/externalLinkN.xml.rels 中 Relationship@Target
    """
    ref_map = {}
    try:
        with zipfile.ZipFile(xlsx_path, 'r') as z:
            rels = ET.fromstring(z.read('xl/_rels/workbook.xml.rels'))
            for rel in rels.findall('{http://schemas.openxmlformats.org/package/2006/relationships}Relationship'):
                if rel.attrib.get('Type','').endswith('/externalLink'):
                    target = rel.attrib.get('Target','')  # e.g., externalLinks/externalLink1.xml
                    m = re.search(r'externalLink(\d+)\.xml', target)
                    if not m:
                        continue
                    num = int(m.group(1))
                    path = ''
                    # 1) 嘗試 externalLinkN.xml 的 externalBookPr@href
                    try:
                        link_xml = z.read(f'xl/{target}')
                        link_tree = ET.fromstring(link_xml)
                        book_elem = link_tree.find('.//{http://schemas.openxmlformats.org/spreadsheetml/2006/main}externalBookPr')
                        if book_elem is not None:
                            path = book_elem.attrib.get('href', '')
                    except Exception:
                        pass
                    # 2) 若仍無，嘗試 externalLinks/_rels/externalLinkN.xml.rels 的 Relationship@Target
                    if not path:
                        try:
                            rels_path = f"xl/externalLinks/_rels/externalLink{num}.xml.rels"
                            if rels_path in z.namelist():
                                link_rels_xml = z.read(rels_path)
                                link_rels = ET.fromstring(link_rels_xml)
                                rel_node = link_rels.find('{http://schemas.openxmlformats.org/package/2006/relationships}Relationship')
                                if rel_node is not None:
                                    path = rel_node.attrib.get('Target','')
                        except Exception:
                            pass
                    ref_map[num] = path or ''
    except (zipfile.BadZipFile, KeyError, ET.ParseError) as e:
        logging.error(f"提取外部參照時發生錯誤: {xlsx_path}, 錯誤: {e}")
    return ref_map

def _normalize_path(p: str) -> str:
    if not p:
        return p
    s = urllib.parse.unquote(p.strip())
    # Handle file: scheme robustly
    try:
        u = urllib.parse.urlparse(s)
        if u.scheme == 'file':
            if u.netloc:  # UNC: file://server/share/path
                path_part = u.path.lstrip('/').replace('/', '\\')
                s = "\\\\" + u.netloc + "\\" + path_part
            else:  # local: file:///C:/path or file:/C:/path or file:\C:\path
                rest = u.path or s[5:]
                rest = rest.lstrip('/\\')
                s = rest.replace('/', '\\')
    except Exception:
        pass
    # Fallback: strip 'file:' prefix crudely if present
    if s.lower().startswith('file:'):
        s = s[5:].lstrip('/\\')
    # normalize backslashes
    s = s.replace('/', '\\')
    # collapse duplicate backslashes but keep UNC prefix
    if s.startswith('\\\\'):
        prefix = '\\'
        t = s[2:]
        while '\\' in t:
            t = t.replace('\\\\', '\\')
        s = '\\' + t
    else:
        while '\\' in s and '\\\\' in s:
            s = s.replace('\\\\', '\\')
    return s


def _excel_external_prefix(norm_path: str, sheet: str) -> str:
    """
    將歸一化路徑與工作表組裝為 Excel 標準外部參照前綴：
    'C:\\dir\\[Workbook.xlsx]Sheet Name'
    注意：整段（目錄 + [檔名] + 工作表）以單引號包裹；工作表名中的單引號需轉義為兩個單引號。
    """
    if not norm_path:
        return None
    # 分割目錄與檔名
    base = os.path.basename(norm_path)
    dir_ = os.path.dirname(norm_path)
    # 若 base 沒有副檔名，原樣處理
    fname = base
    sheet_escaped = (sheet or '').replace("'", "''")
    inside = ''
    if dir_:
        inside = dir_.rstrip('\\') + '\\'
    inside += f"[{fname}]" + sheet_escaped
    return f"'{inside}'"


def pretty_formula(formula, ref_map=None):
    """
    將公式中的外部參照 [n]Sheet! 還原為 'full\\normalized\\path'!Sheet! 的可讀形式。
    同時保留 Excel 語法結構，避免造成假差異。
    """
    if formula is None:
        return None
    
    # 修改：處理 ArrayFormula 物件
    if isinstance(formula, ArrayFormula):
        formula_str = formula.text if hasattr(formula, 'text') else str(formula)
    else:
        formula_str = str(formula)
    
    if ref_map:
        # 1) 直接替換形如 [n]Sheet! 為 'path'!Sheet!
        def repl_path_with_sheet(m):
            n = int(m.group(1))
            sheet = m.group(2)
            raw_path = ref_map.get(n, '')
            norm_path = _normalize_path(raw_path)
            if norm_path:
                prefix = _excel_external_prefix(norm_path, sheet)
                return f"{prefix}!"
            return m.group(0)
        s = re.sub(r"\[(\d+)\]([^!\]]+)!", repl_path_with_sheet, formula_str)
        
        # 2) 對其餘殘留的 [n] 標記（未帶 sheet 名）插入可讀提示
        def repl_annotate(m):
            n = int(m.group(1))
            raw_path = ref_map.get(n, '')
            norm_path = _normalize_path(raw_path)
            if norm_path:
                return f"[外部檔案{n}: {norm_path}]"
            return m.group(0)
        s = re.sub(r"\[(\d+)\]", repl_annotate, s)
        return s
    else:
        return formula_str

def get_cell_formula(cell):
    """
    取得 cell 公式（不論係普通 formula or array formula），一律回傳公式字串
    """
    if cell.data_type == 'f':
        if isinstance(cell.value, ArrayFormula):
            # 修改：返回 ArrayFormula 的實際公式字符串，而不是物件
            return cell.value.text if hasattr(cell.value, 'text') else str(cell.value)
        return cell.value
    return None

def serialize_cell_value(value):
    """
    序列化儲存格值
    """
    if value is None: 
        return None
    if isinstance(value, ArrayFormula): 
        return None
    if isinstance(value, datetime): 
        return value.isoformat()
    if isinstance(value, (int, float, str, bool)): 
        return value
    return str(value)

def get_excel_last_author(path):
    """
    以非鎖定方式讀取 Excel 檔案的最後修改者：
    - 優先從快取副本的 docProps/core.xml 解析 cp:lastModifiedBy（不開啟原檔，不用 openpyxl）。
    - 如遇非常規檔案或解析失敗，才退回以 openpyxl 讀取「快取檔」。
    """
    try:
        # 先複製到本地快取，避免直接打開原始檔案
        local_path = copy_to_cache(path, silent=True)
        if not local_path or not os.path.exists(local_path):
            return None
        try:
            with zipfile.ZipFile(local_path, 'r') as z:
                core_xml = z.read('docProps/core.xml')
                root = ET.fromstring(core_xml)
                ns = {
                    'cp': 'http://schemas.openxmlformats.org/package/2006/metadata/core-properties',
                    'dc': 'http://purl.org/dc/elements/1.1/'
                }
                node = root.find('cp:lastModifiedBy', ns)
                if node is None:
                    node = root.find('dc:lastModifiedBy', ns)  # 極少數模板可能使用 dc
                author = (node.text or '').strip() if node is not None else None
                return author or None
        except (KeyError, zipfile.BadZipFile, ET.ParseError):
            # 結構異常或非 zip 格式（例如舊 xls），退回 openpyxl（仍用本地快取檔）
            pass

        # Fallback：對快取檔使用 openpyxl（不會鎖定原檔）
        try:
            wb = load_workbook(local_path, read_only=True)
            author = wb.properties.lastModifiedBy
            wb.close()
            del wb
            return author
        except Exception as e:
            logging.warning(f"openpyxl 讀取核心屬性失敗: {local_path}, {e}")
            return None

    except FileNotFoundError:
        logging.warning(f"檔案未找到: {path}")
        return None
    except PermissionError:
        logging.error(f"權限不足: {path}")
        return None
    except OSError as e:
        logging.error(f"Excel 檔案讀取 I/O 錯誤: {path}, {e}")
        return None

def safe_load_workbook(path, max_retry=5, delay=0.5, **kwargs):
    """
    安全載入 Excel 檔案，帶重試機制
    """
    last_err = None
    for i in range(max_retry):
        try:
            wb = load_workbook(path, **kwargs)
            return wb
        except PermissionError as e:
            last_err = e
            time.sleep(delay)
        except Exception as e:
            last_err = e
            logging.error(f"載入 Excel 檔案時發生意外錯誤: {path}, 錯誤: {e}")
            break
    raise last_err

def dump_excel_cells_with_timeout(path, show_sheet_detail=True, silent=False):  # noqa: C901
    """
    提取 Excel 檔案中的所有儲存格數據（含公式）
    """
    # 更新全局變數
    settings.current_processing_file = path
    settings.processing_start_time = time.time()
    
    wb = None
    try:
        if not silent: 
            print(f"   📊 檔案大小: {os.path.getsize(path)/(1024*1024):.1f} MB")
        
        local_path = copy_to_cache(path, silent=silent)
        if not local_path or not os.path.exists(local_path):
            if not silent:
                print("   ❌ 無法使用快取副本（嚴格模式下不會讀取原檔），略過此檔案。")
            return None
        
        read_only_mode = True
        if not silent: 
            print(f"   🚀 讀取模式: read_only={read_only_mode}, data_only=False")
        
        wb = safe_load_workbook(local_path, read_only=read_only_mode, data_only=False)
        result = {}
        worksheet_count = len(wb.worksheets)
        
        if not silent and show_sheet_detail: 
            print(f"   工作表數量: {worksheet_count}")
        
        # 解析一次外部參照映射，供 prettify 使用
        ref_map = extract_external_refs(local_path)
        formula_cells_global = 0
        formula_coords_by_sheet = {}

        # 先準備值引擎：polars 或 xml（直接使用全域 settings，避免在函數內重新 import 造成遮蔽）
        value_engine = getattr(settings, 'VALUE_ENGINE', 'polars')
        persist_csv = bool(getattr(settings, 'CSV_PERSIST', False))
        persist_dir = getattr(settings, 'CACHE_FOLDER', None)
        values_by_sheet = {}
        try:
            if value_engine == 'polars':
                # 優先使用 xlsx2csv+polars（若失敗會 fallback 到 polars_xml）
                from utils.value_engines.polars_reader import read_values_from_xlsx_via_polars
                # Debug: print engine/version
                if not silent:
                    try:
                        import polars as _pl
                        import importlib
                        _x2c = importlib.util.find_spec('xlsx2csv') is not None
                        print(f"   [value-engine] POLARS (polars={_pl.__version__}, xlsx2csv={'OK' if _x2c else 'NOT FOUND'}) | persist_csv={persist_csv}")
                    except Exception:
                        print("   [value-engine] POLARS (version info unavailable)")
                try:
                    values_by_sheet = read_values_from_xlsx_via_polars(local_path, persist_csv=persist_csv, persist_dir=persist_dir, sheet_count=len(wb.worksheets))
                except TypeError:
                    # 兼容舊版函數簽名（沒有 sheet_count 參數）
                    values_by_sheet = read_values_from_xlsx_via_polars(local_path, persist_csv=persist_csv, persist_dir=persist_dir)
                # 若 polars 提供的非空值總數為 0，回退到 polars_xml
                try:
                    nonempty_total = sum(len(v or {}) for v in (values_by_sheet or {}).values())
                except Exception:
                    nonempty_total = 0
                if nonempty_total == 0:
                    if not silent:
                        print("   [fallback->polars_xml] reason=polars_nonempty=0")
                    from utils.value_engines.polars_xml_reader import read_values_from_xlsx_via_polars_xml
                    values_by_sheet = read_values_from_xlsx_via_polars_xml(local_path)
                    value_engine = 'polars_xml'
            elif value_engine == 'polars_xml':
                from utils.value_engines.polars_xml_reader import read_values_from_xlsx_via_polars_xml
                if not silent:
                    print("   [value-engine] POLARS_XML (internal XML parser -> values)")
                values_by_sheet = read_values_from_xlsx_via_polars_xml(local_path)
            else:
                from utils.value_engines.xml_reader import read_values_from_xlsx_via_xml
                if not silent:
                    print("   [value-engine] XML parser for values")
                values_by_sheet = read_values_from_xlsx_via_xml(local_path)
        except Exception as e:
            # fallback to xml with diagnostics
            try:
                import sys, importlib.util
                polars_ok = importlib.util.find_spec('polars') is not None
                x2c_ok = importlib.util.find_spec('xlsx2csv') is not None
                if not silent:
                    print(f"   [fallback] Polars/xlsx2csv 不可用，已回退到 XML 值引擎 | python={sys.executable} | polars={'OK' if polars_ok else 'NOT FOUND'} | xlsx2csv={'OK' if x2c_ok else 'NOT FOUND'} | err={e}")
            except Exception:
                pass
            try:
                from utils.value_engines.xml_reader import read_values_from_xlsx_via_xml
                values_by_sheet = read_values_from_xlsx_via_xml(local_path)
            except Exception:
                values_by_sheet = {}
        # 準備 sheet 對齊順序與調試輸出
        try:
            sheet_order = list(values_by_sheet.keys())
            if not silent:
                print(f"   [value-engine] sheet keys from engine: {sheet_order}")
                # 針對 polars_xml 額外輸出每張表的統計與樣本
                if value_engine == 'polars_xml':
                    try:
                        for i, nm in enumerate(sheet_order, start=1):
                            vals = values_by_sheet.get(nm) or {}
                            cells = len(vals)
                            try:
                                nonempty = sum(1 for v in vals.values() if v not in (None, ''))
                            except Exception:
                                nonempty = cells
                            print(f"   [polars_xml] sheet {i} name='{nm}' cells={cells} nonempty={nonempty}")
                            if vals:
                                sample_items = list(vals.items())[:8]
                                sample_str = ', '.join([f"{k}->{repr(v)}" for k, v in sample_items])
                                print(f"   [polars_xml] sample: {sample_str}")
                    except Exception:
                        pass
        except Exception:
            sheet_order = []
        # 若值引擎未能返回任何工作表，直接回退到 XML 值引擎
        if (not values_by_sheet) or (not sheet_order):
            try:
                if not silent:
                    print("   [value-engine] no sheets from POLARS, fallback to XML value engine")
                from utils.value_engines.xml_reader import read_values_from_xlsx_via_xml
                values_by_sheet = read_values_from_xlsx_via_xml(local_path)
                sheet_order = list(values_by_sheet.keys()) if isinstance(values_by_sheet, dict) else []
                if not silent:
                    print(f"   [value-engine] XML sheet keys: {sheet_order}")
            except Exception as _e:
                if not silent:
                    print(f"   [value-engine] XML fallback failed: {_e}")
                values_by_sheet = {}
                sheet_order = []

        per_sheet_formula_provided = {}
        for idx, ws in enumerate(wb.worksheets, 1):
            cell_count = 0
            ws_data = {}
            formula_addrs = []

            # 決定本工作表對應的值引擎 key 與資料
            selected_key = ws.title if ws.title in (values_by_sheet or {}) else None
            if selected_key is None and sheet_order:
                selected_key = list(values_by_sheet.keys())[idx-1] if idx-1 < len(values_by_sheet) else None
                if not silent and selected_key:
                    print(f"   [value-engine] sheet name mismatch: ws.title='{ws.title}' -> fallback to index key='{selected_key}'")
            sheet_vals = (values_by_sheet or {}).get(selected_key, {}) if selected_key else {}
            try:
                p_count = len(sheet_vals)
            except Exception:
                p_count = 0
            if not silent:
                # 列出最多前 50 個 keys，避免超長輸出
                try:
                    keys_list = list(sheet_vals.keys()) if isinstance(sheet_vals, dict) else []
                    show_keys = keys_list[:50]
                    if len(keys_list) > 50:
                        show_keys.append('...')
                except Exception:
                    show_keys = []
                print(f"   [map] ws_index={idx} ws_title='{ws.title}' -> key='{selected_key or ''}' provided={p_count} keys={show_keys}")

            if ws.max_row > 1 or ws.max_column > 1:
                try:
                    # 使用索引安全推導地址，避免 EmptyCell 無 coordinate 造成早退
                    def _col_to_letters(n: int) -> str:
                        s = ''
                        while n > 0:
                            n, r = divmod(n-1, 26)
                            s = chr(65 + r) + s
                        return s
                    for r_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column, values_only=False), start=1):
                        for c_idx, cell in enumerate(row, start=1):
                            addr = f"{_col_to_letters(c_idx)}{r_idx}"
                            # ⚡️ Patch: formula 直接存 cell.formula if present, fallback get_cell_formula
                            try:
                                if hasattr(cell, 'formula') and cell.formula:
                                    fstr = cell.formula
                                else:
                                    fstr = get_cell_formula(cell)
                            except Exception:
                                fstr = None
                            # 對外部參照做正規化展示（還原路徑，解 %20，統一反斜線）
                            if fstr:
                                try:
                                    fstr = pretty_formula(fstr, ref_map=ref_map)
                                except Exception:
                                    pass
                                formula_addrs.append(addr)
                                formula_cells_global += 1
                            # 取值（由值引擎供應）
                            try:
                                vstr = sheet_vals.get(addr)
                            except Exception as _e:
                                if not silent:
                                    print(f"   [read_error] sheet='{ws.title}' addr='{addr}' op='assemble' err={_e}")
                                try:
                                    vstr = serialize_cell_value(getattr(cell, 'value', None))
                                except Exception:
                                    vstr = None
                            if fstr is not None or vstr is not None:
                                # 若值引擎已提供顯示值，直接作為 cached_value 使用，避免後續二次 data_only pass
                                cached_v = vstr if value_engine in ('polars','polars_xml','xml') else None
                                ws_data[addr] = {"formula": fstr, "value": vstr, "cached_value": cached_v}
                                if fstr and (vstr is not None):
                                    per_sheet_formula_provided[selected_key or ws.title] = per_sheet_formula_provided.get(selected_key or ws.title, 0) + 1
                                cell_count += 1
                except Exception as _e:
                    if not silent:
                        print(f"   [read_error] sheet='{ws.title}' op='iterate_rows' err={_e}")
            
            if show_sheet_detail and not silent: 
                print(f"      處理工作表 {idx}/{worksheet_count}: {ws.title}（{cell_count} 有資料 cell）")
            
            if ws_data: 
                result[ws.title] = ws_data
            if formula_addrs:
                formula_coords_by_sheet[ws.title] = formula_addrs

        # Phase 2：可選的 cached value 比對（僅對公式格），避免外部參照刷新導致假變更
        try:
            if getattr(settings, 'ENABLE_FORMULA_VALUE_CHECK', False) and formula_cells_global > 0:
                # 若值引擎已提供 cached_value，則無需再做第二次 data_only pass
                provided = 0
                for sheet_name, coords in formula_coords_by_sheet.items():
                    for addr in coords:
                        if sheet_name in result and addr in result[sheet_name] and result[sheet_name][addr].get('cached_value') is not None:
                            provided += 1
                need_data_only = (provided == 0)
                if not need_data_only:
                    if not silent:
                        # 按工作表顯示提供情況
                        try:
                            per_sheet_counts = {}
                            for sname, coords in formula_coords_by_sheet.items():
                                cnt = 0
                                for addr in coords:
                                    if sname in result and addr in result[sname] and result[sname][addr].get('cached_value') is not None:
                                        cnt += 1
                                per_sheet_counts[sname] = cnt
                            print(f"   [phase2] provided_from_value_engine total={provided} per_sheet={per_sheet_counts}")
                        except Exception:
                            print(f"   [phase2] 已由值引擎提供 cached value（{provided} 格），略過 openpyxl data_only 二次讀取。")
                else:
                    cap = int(getattr(settings, 'MAX_FORMULA_VALUE_CELLS', 50000))
                    if formula_cells_global > cap:
                        if not silent:
                            print(f"   ⏩ 公式格數量 {formula_cells_global} 超過上限 {cap}，略過值比對。")
                    else:
                        if not silent:
                            print(f"   讀取公式儲存格的 cached value（共 {formula_cells_global} 格）…")
                        wb_values = safe_load_workbook(local_path, read_only=True, data_only=True)
                        try:
                            for sheet_name, coords in formula_coords_by_sheet.items():
                                if sheet_name not in wb_values.sheetnames:
                                    continue
                                ws2 = wb_values[sheet_name]
                                for addr in coords:
                                    try:
                                        val = ws2[addr].value
                                    except Exception:
                                        val = None
                                    sval = serialize_cell_value(val)
                                    if sheet_name in result and addr in result[sheet_name]:
                                        result[sheet_name][addr]['cached_value'] = sval
                        finally:
                            try:
                                wb_values.close()
                            except Exception:
                                pass
        except Exception as e:
            logging.warning(f"讀取 cached value 失敗：{e}")
        
        wb.close()
        wb = None
        
        if not silent and show_sheet_detail: 
            print(f"   ✅ Excel 讀取完成")
        
        return result
        
    except Exception as e:
        if not silent: 
            logging.error(f"Excel 讀取失敗: {e}")
        return None
    finally:
        if wb: 
            wb.close()
            del wb
        
        # 重置全局變數
        settings.current_processing_file = None
        settings.processing_start_time = None

def hash_excel_content(cells_dict):
    """
    計算 Excel 內容的雜湊值
    """
    if cells_dict is None: 
        return None
    
    try:
        content_str = json.dumps(cells_dict, sort_keys=True, ensure_ascii=False)
        return hashlib.md5(content_str.encode('utf-8')).hexdigest()
    except (TypeError, json.JSONEncodeError) as e:
        logging.error(f"計算 Excel 內容雜湊值失敗: {e}")
        return None