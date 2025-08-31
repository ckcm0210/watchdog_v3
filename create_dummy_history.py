# -*- coding: utf-8 -*-
"""
建立示範 Git 倉庫與提交（完全自動）：
- 建立/初始化 excel_git_repo
- 以 openpyxl 產生有效的 sample.xlsx，並提交兩個版本（B2=Version1 → Version2）
- 同時產生 sample.cells.json（純文字），便於在 Git 內做差異比較
使用：
  雙擊 run_create_dummy_history.bat
或：
  python create_dummy_history.py
"""
import os
from datetime import datetime

REPO_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'excel_git_repo')


def ensure_repo():
    from git import Repo
    os.makedirs(REPO_PATH, exist_ok=True)
    try:
        return Repo(REPO_PATH)
    except Exception:
        return Repo.init(REPO_PATH)


def write_excel(path: str, value: str):
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = '工作表1'
    ws['A1'] = 'Hello'
    ws['B2'] = value
    ws['C3'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    wb.save(path)


def write_json(path: str, value: str):
    import json
    payload = {
        'sheet': '工作表1',
        'cells': {
            'A1': {'value': 'Hello', 'formula': None},
            'B2': {'value': value, 'formula': None},
        },
        'timestamp': datetime.now().isoformat()
    }
    with open(path, 'w', encoding='utf-8') as f:
        json.dump(payload, f, ensure_ascii=False, indent=2)


def commit(repo, message: str, filenames):
    repo.index.add(filenames)
    repo.index.commit(message)


def main():
    from git import Repo
    repo = ensure_repo()
    sample_xlsx = os.path.join(REPO_PATH, 'sample.xlsx')
    sample_json = os.path.join(REPO_PATH, 'sample.cells.json')

    # 版本 1
    write_excel(sample_xlsx, 'Version1')
    write_json(sample_json, 'Version1')
    commit(repo, '新增 sample.xlsx（B2=Version1）與 sample.cells.json', ['sample.xlsx', 'sample.cells.json'])

    # 版本 2
    write_excel(sample_xlsx, 'Version2')
    write_json(sample_json, 'Version2')
    commit(repo, '更新 sample.xlsx（B2=Version2）與 sample.cells.json', ['sample.xlsx', 'sample.cells.json'])

    print('✅ 測試倉庫已準備好：', REPO_PATH)
    print('   - 建立了 sample.xlsx 與 sample.cells.json 的兩個版本提交')
    print('   - 你可以執行 run_git_viewer.bat 來瀏覽歷史')


if __name__ == '__main__':
    main()
