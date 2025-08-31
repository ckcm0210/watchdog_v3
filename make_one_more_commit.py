# -*- coding: utf-8 -*-
"""
一鍵在 excel_git_repo 產生第三個版本（Version3）：
- 更新 sample.cells.json 的 B2 值為 "Version3"
- 同步更新 sample.xlsx 的 B2 為 "Version3"
- 自動 git add + commit
用法：
  雙擊 run_make_one_more_commit.bat 或 python make_one_more_commit.py
"""
import os
import json
from datetime import datetime

REPO_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'excel_git_repo')


def ensure_repo():
    from git import Repo
    os.makedirs(REPO_PATH, exist_ok=True)
    try:
        return Repo(REPO_PATH)
    except Exception:
        return Repo.init(REPO_PATH)


def update_cells_json(path: str):
    data = {}
    try:
        with open(path, 'r', encoding='utf-8') as f:
            data = json.load(f)
    except Exception:
        data = {}
    # 安全更新結構
    if 'cells' not in data or not isinstance(data['cells'], dict):
        data['cells'] = {}
    if 'B2' not in data['cells'] or not isinstance(data['cells']['B2'], dict):
        data['cells']['B2'] = {}
    data['cells']['B2']['value'] = 'Version3'
    data['cells']['B2']['formula'] = None
    data['timestamp'] = datetime.now().isoformat()
    with open(path, 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False, indent=2)


def update_xlsx(path: str):
    try:
        from openpyxl import load_workbook
        wb = load_workbook(path)
        ws = wb.active
        ws['B2'] = 'Version3'
        wb.save(path)
        try:
            wb.close()
        except Exception:
            pass
    except Exception:
        # 若 openpyxl 或檔案讀寫失敗，略過 xlsx 更新
        pass


def main():
    from git import Repo
    repo = ensure_repo()
    sample_json = os.path.join(REPO_PATH, 'sample.cells.json')
    sample_xlsx = os.path.join(REPO_PATH, 'sample.xlsx')
    if not os.path.exists(sample_json) and not os.path.exists(sample_xlsx):
        print('找不到 sample 檔案，請先執行 run_create_dummy_history.bat')
        return
    if os.path.exists(sample_json):
        update_cells_json(sample_json)
    if os.path.exists(sample_xlsx):
        update_xlsx(sample_xlsx)
    # commit
    files = []
    if os.path.exists(sample_json):
        files.append('sample.cells.json')
    if os.path.exists(sample_xlsx):
        files.append('sample.xlsx')
    if not files:
        print('沒有可提交的檔案。')
        return
    repo.index.add(files)
    repo.index.commit('更新 sample 至 Version3（B2=Version3）')
    print('✅ 已建立 Version3 提交。你可以在 viewer 中選擇兩個版本比較差異。')


if __name__ == '__main__':
    main()
